import requests
import time
from datasets import Dataset
import pandas as pd
import os
import sys
import logging
from pydantic import BaseModel, Field
import asyncio
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from dotenv import load_dotenv
import re
from typing import Dict, Any, List
import glob


sys.path.append("./app")
from backend.llm.gemini_client import GeminiClient
from backend.db.neo4j_client import Neo4jClient
from backend.db.vector_db import VectorDBClient
from backend.llm.ollama_client import OllamaClient
from qdrant_client import QdrantClient
from backend.core.retrieval.kg_query_processor import run_query


logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("multi_eval.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class EvaluationResult(BaseModel):
    is_correct: int = Field()

class TranslationResult(BaseModel):
    translated_text: str = Field()

class GeminiRateLimiter:
    """Rate limiter for Gemini API calls (10 requests per minute)"""
    def __init__(self, max_requests_per_minute=10):
        self.max_requests = max_requests_per_minute
        self.requests = []
        self.lock = asyncio.Lock()
    
    async def wait_if_needed(self):
        """Wait if we're approaching rate limit"""
        async with self.lock:
            now = time.time()
            # Remove requests older than 1 minute
            self.requests = [req_time for req_time in self.requests if now - req_time < 60]
            
            if len(self.requests) >= self.max_requests:
                # Calculate how long to wait
                oldest_request = min(self.requests)
                wait_time = 61 - (now - oldest_request)
                if wait_time > 0:
                    logger.info(f"Rate limit approaching, waiting {wait_time:.1f} seconds...")
                    await asyncio.sleep(wait_time)
                    # Clean up again after waiting
                    now = time.time()
                    self.requests = [req_time for req_time in self.requests if now - req_time < 60]
            
            # Record this request
            self.requests.append(now)

class GeminiKeyManager:
    def __init__(
        self,
        current_key_index: int = 1,
        max_keys: int = 6,
        key_pattern: str = "GEMINI_API_KEY_{}"
    ):
        self.current_key_index = current_key_index
        self.max_keys = max_keys
        self.key_pattern = key_pattern
        self.logger = logging.getLogger(__name__)
        self.rate_limiter = GeminiRateLimiter()
    
    def get_current_key(self) -> str:
        key_name = self.key_pattern.format(self.current_key_index)
        api_key = os.getenv(key_name)
        #print(f"Using API key: {key_name} - {api_key is not None}")
        
        if not api_key:
            self.logger.warning(f"API key {key_name} not found in environment variables")
            return None
            
        return api_key
    
    def rotate_key(self) -> str:
        for _ in range(self.max_keys):
            self.current_key_index = (self.current_key_index % self.max_keys) + 1
            
            key_name = self.key_pattern.format(self.current_key_index)
            api_key = os.getenv(key_name)
            
            if api_key:
                self.logger.info(f"Rotated to API key {key_name}")
                # Reset rate limiter for new key
                self.rate_limiter = GeminiRateLimiter()
                return api_key
        
        self.logger.error("No valid API keys found after trying all options")
        return None

class ClientManager:
    """Singleton để quản lý các client connections"""
    _instance = None
    
    def __new__(cls):
        if cls._instance is None:
            cls._instance = super(ClientManager, cls).__new__(cls)
            cls._instance.initialized = False
        return cls._instance
    
    async def initialize(self, gemini_key_manager):
        """Initialize all backend clients"""
        if self.initialized:
            return self.clients
            
        load_dotenv()
        
        # Initialize Neo4j client
        neo4j_uri = os.getenv('NEO4J_URI')
        neo4j_username = os.getenv('NEO4J_USERNAME')
        neo4j_password = os.getenv('NEO4J_PASSWORD')
        
        self.neo4j_client = Neo4jClient(
            uri=neo4j_uri,
            username=neo4j_username,
            password=neo4j_password
        )
        
        # Initialize Vector DB client
        qdrant_host = os.getenv("QDRANT_HOST", "localhost")
        qdrant_port = int(os.getenv("QDRANT_PORT", "6333"))
        
        self.vector_db_client = VectorDBClient(
            host=qdrant_host,
            port=qdrant_port
        )
        
        # Initialize Qdrant client
        self.qdrant_client = QdrantClient(
            host=qdrant_host, 
            port=qdrant_port
        )
        
        # Initialize Ollama client
        ollama_host = os.getenv("OLLAMA_HOST", "http://localhost:11434")
        embedding_model = os.getenv("OLLAMA_EMBEDDING_MODEL", "mxbai-embed-large")
        
        self.ollama_client = OllamaClient(
            host=ollama_host,
            embedding_model=embedding_model
        )
        
        # Initialize Gemini client with key manager
        gemini_api_key = gemini_key_manager.get_current_key()
        self.gemini_client = GeminiClient(
            api_key=gemini_api_key, 
            model_name="gemini-2.0-flash"
        )
        self.gemini_key_manager = gemini_key_manager
        
        # Verify connectivity
        if not await self.neo4j_client.verify_connectivity():
            logger.error("Failed to connect to Neo4j database")
            raise Exception("Neo4j connection failed")
            
        logger.info("All backend clients initialized successfully")
        
        self.clients = {
            "neo4j_client": self.neo4j_client,
            "vector_db_client": self.vector_db_client,
            "ollama_client": self.ollama_client,
            "gemini_client": self.gemini_client,
            "qdrant_client": self.qdrant_client
        }
        
        self.initialized = True
        return self.clients
    
    def rotate_gemini_key(self):
        """Rotate Gemini API key when needed"""
        new_key = self.gemini_key_manager.rotate_key()
        if new_key:
            self.gemini_client = GeminiClient(
                api_key=new_key, 
                model_name="gemini-2.0-flash"
            )
            self.clients["gemini_client"] = self.gemini_client
            return self.gemini_client
        return None
    
    async def close(self):
        """Close all client connections"""
        if hasattr(self, 'neo4j_client'):
            await self.neo4j_client.close()
        self.initialized = False

def load_csv_datasets(csv_paths: List[str]) -> pd.DataFrame:
    """
    Load and combine multiple CSV files into a single DataFrame
    """
    all_dataframes = []
    
    for csv_path in csv_paths:
        if not os.path.exists(csv_path):
            logger.warning(f"CSV file not found: {csv_path}")
            continue
            
        try:
            logger.info(f"Loading CSV file: {csv_path}")
            df = pd.read_csv(csv_path)
            
            # Verify expected columns
            expected_columns = ['question', 'answer']
            if not all(col in df.columns for col in expected_columns):
                logger.error(f"CSV file {csv_path} missing expected columns {expected_columns}")
                logger.error(f"Found columns: {list(df.columns)}")
                continue
            
            # Add source file info
            df['source_file'] = os.path.basename(csv_path)
            all_dataframes.append(df)
            logger.info(f"Loaded {len(df)} questions from {csv_path}")
            
        except Exception as e:
            logger.error(f"Error loading CSV file {csv_path}: {str(e)}")
            continue
    
    if not all_dataframes:
        raise ValueError("No valid CSV files could be loaded")
    
    # Combine all dataframes
    combined_df = pd.concat(all_dataframes, ignore_index=True)
    logger.info(f"Combined dataset: {len(combined_df)} total questions from {len(all_dataframes)} files")
    
    return combined_df



async def translate_chinese_to_english(gemini_client, key_manager, text, max_retries=3):
    """
    Translate Chinese text to English using Gemini client
    """
    translation_prompt = f"""
    Please translate the following Chinese medical text to English. 
    This is a multiple-choice question about diabetes or endocrinology.
    
    Instructions:
    - Translate the question and all answer options accurately
    - Maintain the structure (Question:, Option A:, Option B:, etc.)
    - Use proper medical terminology
    - Provide only the English translation without any additional explanation
    
    Chinese text:
    {text}
    """
    
    for attempt in range(max_retries):
        try:
            # Wait for rate limiting
            # await key_manager.rate_limiter.wait_if_needed()
            
            logger.debug(f"Translating text (attempt {attempt + 1}/{max_retries})")
            
            response = await gemini_client.generate(
                prompt=translation_prompt,
                format=TranslationResult
            )
            
            if isinstance(response, list) and len(response) > 0:
                translated_text = response[0].translated_text
                logger.debug(f"Translation successful: {text[:50]}... -> {translated_text[:50]}...")
                return translated_text
            else:
                logger.warning(f"No translation received on attempt {attempt + 1}")
                
        except Exception as e:
            error_str = str(e).lower()
            logger.error(f"Translation error on attempt {attempt + 1}/{max_retries}: {error_str}")
            
            # Check if it's a rate limit or quota error
            if any(keyword in error_str for keyword in ["resource_exhausted", "rate limit", "429", "quota"]):
                new_key = key_manager.rotate_key()
                if new_key:
                    gemini_client.api_key = new_key
                    os.environ["GEMINI_API_KEY"] = new_key
                    logger.info(f"Switched to API key index {key_manager.current_key_index} for translation")
                else:
                    logger.error("No more API keys available for translation")
                    if attempt == max_retries - 1:
                        break
            
            if attempt < max_retries - 1:
                await asyncio.sleep(3)  # Wait longer for translation retries
    
    logger.error(f"Failed to translate text after {max_retries} attempts: {text[:100]}...")
    return text  # Return original text if translation fails

async def ask_question_direct(question, clients, key_manager, response_type="concise", max_retries=3):
    """
    Directly use backend services instead of API calls
    """
    guided_question = f"""
    This is a medical question about diabetes or endocrinology. Please read the question carefully and provide your answer.

    QUESTION TYPES AND INSTRUCTIONS:
    - Determine the type of question besed on the content
    - For multiple-choice questions (A, B, C, D, E): Select the most appropriate option(s)
    - For yes/no questions: Answer "Yes" or "No"
    - For questions requiring multiple answers: List all correct options
    - For open-ended questions: Provide a concise answer

    RESPONSE FORMAT:
    Please provide your response in the following format:
    
    ANSWER: [Your answer - option letter(s), yes/no, or brief response]
    
    EXPLANATION: [Brief explanation (1-2 sentences) justifying your answer choice]

    QUESTION:
    {question}
    """
    
    conversation_history = []
    
    for attempt in range(max_retries):
        try:
            logger.debug(f"Processing question (attempt {attempt + 1}/{max_retries})")
            
            result = await run_query(
                query=guided_question,
                conversation_history=conversation_history,
                clients=clients,
                grounding=False,
                language="English"
            )
            
            if result and "response" in result:
                return result["response"]
            else:
                logger.warning(f"No response received on attempt {attempt + 1}")
                
        except Exception as e:
            error_str = str(e).lower()
            logger.error(f"Error on attempt {attempt + 1}/{max_retries}: {error_str}")
            
            # Check if it's a rate limit or quota error
            if any(keyword in error_str for keyword in ["resource_exhausted", "rate limit", "429", "quota"]):
                # Try to rotate the Gemini API key
                client_manager = ClientManager()
                new_client = client_manager.rotate_gemini_key()
                if new_client:
                    clients["gemini_client"] = new_client
                    logger.info(f"Rotated to new Gemini API key (index: {key_manager.current_key_index})")
                    await asyncio.sleep(3)  # Wait before retry
                else:
                    logger.error("No more valid Gemini API keys available")
                    if attempt == max_retries - 1:
                        break
            else:
                # For other errors, wait a bit and retry
                await asyncio.sleep(2)
                
        if attempt < max_retries - 1:
            await asyncio.sleep(3)  # Wait between retries
    
    logger.error(f"Failed to get response after {max_retries} attempts")
    return "Error: Failed to get response from backend services"

async def evaluate_answer(gemini_client, key_manager, question, correct_answer, model_answer):
    """
    Evaluate if model's answer matches the correct answer
    """
    prompt = f"""
    You are an evaluation system for cross-language medical question answers.
    
    Context:
    - The original question was in Chinese and has been translated to English
    - The correct answer is in Chinese format
    - The model's answer is in English
    
    Question (English translation):
    {question}
    
    Correct answer (Chinese): {correct_answer}
    
    Model's answer (English): {model_answer}
    
    Your task is to determine if the model's English answer corresponds to the correct Chinese answer.
    
    EVALUATION RULES:
    1. Extract the option letter from the Chinese correct answer (usually the first character: A, B, C, D, E)
    2. Look for the same option letter in the model's English answer
    3. Handle different formats:
       - Chinese format: "A：选项内容" or "A: 选项内容"
       - English format: "ANSWER: A" or "A)" or just "A"
    4. Focus on the option letter match, not the content language
    5. Be flexible with formatting - the model might include explanations
    
    EXAMPLES:
    - Chinese correct: "A：血浆皮质醇测定", Model: "ANSWER: A" → return 1
    - Chinese correct: "E：促肾上腺皮质激素试验", Model: "The answer is E" → return 1  
    - Chinese correct: "B：尿游离皮质醇", Model: "ANSWER: A" → return 0
    - Chinese correct: "C：测定", Model: "C) This option is correct" → return 1
    
    IMPORTANT: Focus only on whether the option letters match, regardless of language differences.
    
    Return 1 if the option letters match, 0 if they don't match.
    Do not provide explanations, just the binary evaluation result.
    """
    
    max_retries = 3
    
    for attempt in range(max_retries):
        try:
            # Wait for rate limiting
            # await key_manager.rate_limiter.wait_if_needed()
            
            response = await gemini_client.generate(
                prompt=prompt,
                format=EvaluationResult
            )
            
            if isinstance(response, list) and len(response) > 0:
                result = response[0].is_correct
            else:
                result = 0                
            return result
            
        except Exception as e:
            logger.error(f"Error during evaluation (attempt {attempt+1}/{max_retries}): {str(e)}")
            
            if attempt < max_retries - 1:
                new_key = key_manager.rotate_key()
                if new_key:
                    gemini_client.api_key = new_key
                    os.environ["GEMINI_API_KEY"] = new_key
                    logger.info(f"Switched to API key index {key_manager.current_key_index} for next attempt")
                else:
                    logger.error("No more API keys available to try")
                    break
                
                await asyncio.sleep(3)  # Longer wait for evaluation retries
    
    logger.warning("All evaluation attempts failed")
    return 0

def create_excel_with_formatting(df, output_path):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Evaluation Results')
    
    workbook = writer.book
    worksheet = writer.sheets['Evaluation Results']
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    correct_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    incorrect_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Adjust column widths
    column_widths = {
        'A': 15,  # source_file
        'B': 60,  # original_question (Chinese)
        'C': 60,  # translated_question (English)
        'D': 30,  # correct_answer
        'E': 30,  # model_answer
        'F': 15   # is_correct
    }
    
    for col_letter, width in column_widths.items():
        worksheet.column_dimensions[col_letter].width = width
    
    # Format headers
    for col_num, column_title in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        
    # Color code the is_correct column
    if 'is_correct' in df.columns:
        is_correct_col = df.columns.get_loc('is_correct') + 1
        for row_num, value in enumerate(df['is_correct'], 2):
            cell = worksheet.cell(row=row_num, column=is_correct_col)
            if value == 1:
                cell.fill = correct_fill
                cell.value = "1"
            else:
                cell.fill = incorrect_fill
                cell.value = "0"
    
    writer.close()
    
    logger.info(f"Excel file created at {output_path} with formatting")

async def process_dataset(df, clients, gemini_client, key_manager, num_samples=None, batch_size=100, save_prefix='diabetes_mc_results'):
    results = []
    
    total_examples = len(df)
    examples_to_process = min(total_examples, num_samples) if num_samples else total_examples
    
    logger.info(f"Starting processing of {examples_to_process} questions")
    
    # Create output directories
    now = datetime.now()
    date_str = now.strftime("%Y_%m_%d")
    time_str = now.strftime("%H_%M_%S")
    
    batch_dir = f"eval/batches/{date_str}/{time_str}"
    os.makedirs(batch_dir, exist_ok=True)
    
    for i in range(examples_to_process):
        row = df.iloc[i]
        original_question = row['question']
        original_answer = row['answer']
        source_file = row.get('source_file', 'unknown')
        
        logger.info(f"Processing question {i+1}/{examples_to_process} from {source_file}")
        logger.debug(f"Original question: {original_question[:100]}...")
        logger.debug(f"Original answer: {original_answer}")
        
        # Translate Chinese question to English
        logger.debug("Translating question from Chinese to English...")
        translated_question = await translate_chinese_to_english(
            gemini_client, 
            key_manager, 
            original_question
        )
        
        # Use translated question for processing
        model_answer = await ask_question_direct(
            question=translated_question, 
            clients=clients, 
            key_manager=key_manager,
            response_type="concise"
        )
        logger.debug(f"Model answered: {model_answer}")
        
        # Evaluate the answer
        is_correct = await evaluate_answer(
            gemini_client, 
            key_manager, 
            translated_question, 
            original_answer, 
            model_answer
        )
        
        results.append({
            'source_file': source_file,
            'original_question': original_question,      
            'translated_question': translated_question,  
            'correct_answer': original_answer,           
            'model_answer': model_answer,
            'is_correct': is_correct
        })
        
        # Save batch periodically
        if (i + 1) % batch_size == 0 or i == examples_to_process - 1:
            batch_number = (i + 1) // batch_size if (i + 1) % batch_size == 0 else ((i + 1) // batch_size) + 1
            batch_df = pd.DataFrame(results)
            batch_filename = f"{batch_dir}/{save_prefix}_batch_{batch_number}.xlsx"
            
            create_excel_with_formatting(batch_df, batch_filename)
            logger.info(f"Saved batch {batch_number} with {len(batch_df)} questions to {batch_filename}")
    
    logger.info(f"Completed processing all {examples_to_process} questions")
    return pd.DataFrame(results)

async def main():
    logger.info("Starting evaluation process with CSV input and Chinese-English translation")
    
    load_dotenv()
    
    csv_paths = [
        "/home/hung/Documents/hung/code/KG_Hung/KGChat/eval/diabetica/ZhiCheng_MCQ_A1.csv",
        "/home/hung/Documents/hung/code/KG_Hung/KGChat/eval/diabetica/ZhiCheng_MCQ_A2.csv"
    ]
    
    key_manager = GeminiKeyManager(current_key_index=1, max_keys=6)
    client_manager = ClientManager()
    
    try:
        clients = await client_manager.initialize(key_manager)
        
        api_key = key_manager.get_current_key()
        if not api_key:
            logger.error("No valid Gemini API key found in environment variables")
            return
        
        gemini_client = GeminiClient(api_key=api_key, model_name="gemini-2.0-flash")
        logger.info(f"Initialized evaluation Gemini client with API key index {key_manager.current_key_index}")
        
        logger.info("Loading CSV datasets...")
        try:
            combined_df = load_csv_datasets(csv_paths)
            logger.info(f"Dataset loaded successfully with {len(combined_df)} questions")
            
        except Exception as e:
            logger.error(f"Failed to load CSV datasets: {str(e)}")
            return
        
        now = datetime.now()
        date_str = now.strftime("%Y_%m_%d")
        time_str = now.strftime("%H_%M_%S")
        
        output_dir = f"eval/multiple_choice/{date_str}/{time_str}"
        os.makedirs(output_dir, exist_ok=True)
        
        num_samples = None  
        if num_samples:
            logger.info(f"Will process {num_samples} samples for testing")
        else:
            logger.info("Processing all samples in the dataset")
        
        results_df = await process_dataset(
            df=combined_df, 
            clients=clients,
            gemini_client=gemini_client,
            key_manager=key_manager,
            num_samples=num_samples,
            batch_size=50,
            save_prefix='zhicheng_mcq_results'
        )
        
        # Create final Excel file with formatting
        final_output_path = f"{output_dir}/zhicheng_mcq_results_complete.xlsx"
        create_excel_with_formatting(results_df, final_output_path)
        logger.info(f"Saved complete results to {final_output_path}")
        
        # Calculate and log summary statistics
        total_questions = len(results_df)
        correct_count = results_df['is_correct'].sum()
        accuracy = (correct_count / total_questions) * 100 if total_questions > 0 else 0
        
        # Statistics by source file
        source_stats = results_df.groupby('source_file').agg({
            'is_correct': ['count', 'sum']
        }).round(2)
        source_stats.columns = ['Total', 'Correct']
        source_stats['Accuracy'] = (source_stats['Correct'] / source_stats['Total'] * 100).round(2)
        
        logger.info("\nResults Summary:")
        logger.info(f"Total questions processed: {total_questions}")
        logger.info(f"Correct answers: {correct_count} ({accuracy:.2f}%)")
        logger.info("\nResults by source file:")
        for source_file, stats in source_stats.iterrows():
            logger.info(f"{source_file}: {stats['Correct']}/{stats['Total']} ({stats['Accuracy']:.1f}%)")
        
        # Create summary sheet
        summary_df = pd.DataFrame([{
            'Total Questions': total_questions,
            'Correct Answers': correct_count,
            'Accuracy': f"{accuracy:.2f}%",
            'Date': pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
            'Translation Enabled': True,
            'Input Format': 'CSV (Chinese MCQ)',
            'Files Processed': len(csv_paths)
        }])
        
        summary_file = f"{output_dir}/zhicheng_mcq_results_summary.xlsx"
        summary_df.to_excel(summary_file, index=False)
        logger.info(f"Saved summary statistics to {summary_file}")
        
        # Save detailed source statistics
        source_stats_file = f"{output_dir}/zhicheng_source_statistics.xlsx"
        source_stats.to_excel(source_stats_file)
        logger.info(f"Saved source statistics to {source_stats_file}")
        
    except Exception as e:
        logger.error(f"Error in main process: {str(e)}", exc_info=True)
    finally:
        # Clean up connections
        await client_manager.close()
        logger.info("Closed all backend connections")

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except Exception as e:
        logger.error(f"Unhandled exception: {str(e)}", exc_info=True)